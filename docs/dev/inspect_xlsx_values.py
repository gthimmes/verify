"""Look at the distinct values in Type / Priority + a couple of full Steps/Expected payloads."""

import sys
from collections import Counter
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path, data_only=True, read_only=True)

priorities = Counter()
types = Counter()
folders = Counter()
multistep_examples = []
folder_examples = set()

for sheet_name in wb.sheetnames:
    if sheet_name == "Summary":
        continue
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = ["" if h is None else str(h) for h in header]
    idx = {h: i for i, h in enumerate(header)}
    for row in rows:
        if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        def cell(name):
            i = idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]
        p = cell("Priority")
        if p is not None:
            priorities[p] += 1
        t = cell("Type")
        if t is not None:
            types[t] += 1
        f = cell("Folder")
        if isinstance(f, str):
            folder_examples.add(f.strip())
        steps = cell("Steps")
        if isinstance(steps, str) and steps.count("[") >= 2 and len(multistep_examples) < 3:
            multistep_examples.append((sheet_name, cell("Title"), steps, cell("Expected result")))

print("=== Type distribution ===")
for k, n in types.most_common():
    print(f"  {k!r:30} ×{n}")
print()
print("=== Priority distribution (raw values) ===")
for k, n in priorities.most_common():
    print(f"  {k!r:10} ×{n}")
print()
print(f"=== {len(folder_examples)} distinct folder paths (first 10) ===")
for f in sorted(folder_examples)[:10]:
    print(f"  {f}")
print()
print("=== Multi-step examples (first 3) ===")
for sheet, title, steps, expected in multistep_examples:
    print(f"--- [{sheet}] {title}")
    print(f"  Steps:")
    for line in steps.splitlines()[:12]:
        print(f"    {line!r}")
    print(f"  Expected:")
    for line in (expected or "").splitlines()[:12]:
        print(f"    {line!r}")
    print()
