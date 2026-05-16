"""Inspect a Testiny xlsx export and report which columns carry data."""

import sys
from collections import Counter
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path, data_only=True, read_only=True)

print(f"=== {path} ===\n")
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        print(f"[sheet: {sheet_name}] empty\n")
        continue

    header = ["" if h is None else str(h) for h in header]
    n_cols = len(header)
    nonempty = [0] * n_cols
    examples = [None] * n_cols
    types = [Counter() for _ in range(n_cols)]
    total = 0

    for row in rows:
        if row is None:
            continue
        # pad/truncate to header length
        cells = list(row) + [None] * (n_cols - len(row)) if len(row) < n_cols else list(row)[:n_cols]
        # skip wholly-empty rows
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in cells):
            continue
        total += 1
        for i, c in enumerate(cells):
            if c is None or (isinstance(c, str) and c.strip() == ""):
                continue
            nonempty[i] += 1
            types[i][type(c).__name__] += 1
            if examples[i] is None:
                s = str(c).replace("\n", " ⏎ ").strip()
                if len(s) > 140:
                    s = s[:140] + "…"
                examples[i] = s

    print(f"=== sheet: {sheet_name} ({total} data rows, {n_cols} columns) ===")
    print()
    for i, name in enumerate(header):
        if nonempty[i] == 0:
            continue
        type_str = ", ".join(f"{k}×{v}" for k, v in types[i].most_common())
        print(f"  [{i:2}] {name!r:40} non-empty={nonempty[i]:>4}/{total}  types=({type_str})")
        if examples[i]:
            print(f"        e.g. {examples[i]}")
    print()
    print("--- empty/no-data columns ---")
    empties = [(i, h) for i, h in enumerate(header) if nonempty[i] == 0]
    if empties:
        for i, h in empties:
            print(f"  [{i:2}] {h!r}")
    else:
        print("  (none)")
    print()
