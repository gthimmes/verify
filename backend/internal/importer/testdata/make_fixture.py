"""Generate a minimal Testiny-shaped xlsx fixture for the importer tests.

The fixture is **entirely synthetic** — no real test cases, no employee
names, no customer-identifying paths.  Folder names mirror the *shape* of
a Testiny export (deep `>`-separated paths, a `DEPRECATED` archive root,
a `CRITICAL CHANGES` cross-cut) without using any real project's names.

Re-run with:  python make_fixture.py
The output goes to fixture.xlsx in this directory.
"""
import datetime as dt
from openpyxl import Workbook

HEADER = [
    "Folder", "Test case ID", "Title", "Owner", "Created at", "Created by",
    "Modified at", "Modified by", "Precondition", "Steps", "Expected result",
    "Folder description", "Priority", "Type", "Requirements", "Attachments",
]


def add_sheet(wb, name, rows):
    ws = wb.create_sheet(title=name)
    ws.append(HEADER)
    for r in rows:
        ws.append([r.get(h, None) for h in HEADER])


def main():
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    # Summary sheet — should be skipped by the reader
    summary = wb.create_sheet("Summary")
    summary.append(["Total"])
    summary.append([1])

    # Module A — single + multi-step cases
    add_sheet(wb, "Module A", [
        {
            "Folder": "Demo Project > Module A",
            "Test case ID": 1001,
            "Title": "Open module landing page",
            "Owner": "Alice",
            "Created at": dt.datetime(2025, 9, 26, 15, 26, 27),
            "Created by": "Alice",
            "Modified at": dt.datetime(2025, 9, 26, 15, 26, 27),
            "Modified by": "Alice",
            "Precondition": "User is logged in",
            "Steps": "[1] 1. Navigate to /module-a\n2. Wait for page",
            "Expected result": "[1] The module landing page renders",
            "Type": "FUNCTIONAL",
        },
        {
            "Folder": "Demo Project > Module A > Drafts",
            "Test case ID": 1002,
            "Title": "Draft creation across roles",
            "Owner": "Bob",
            "Created at": dt.datetime(2025, 9, 26, 16, 0, 0),
            "Created by": "Bob",
            "Modified at": dt.datetime(2025, 9, 26, 16, 0, 0),
            "Modified by": "Bob",
            "Precondition": "Test users seeded",
            "Steps": (
                "[1] 1. Sign in as admin\n2. Create draft\n3. Verify status\n\n"
                "[2] 1. Sign in as manager\n2. Create draft\n3. Verify status"
            ),
            "Expected result": "[1] Draft is created\n\n[2] Draft is created",
            "Type": "REGRESSION",
            "Priority": 0,
        },
    ])

    # Module B — exercises Acceptance + Compatibility + Other types
    add_sheet(wb, "Module B", [
        {
            "Folder": "Demo Project > Module B > Settings",
            "Test case ID": 2001,
            "Title": "Settings page loads",
            "Owner": "Carol",
            "Created at": dt.datetime(2025, 9, 27, 9, 0, 0),
            "Created by": "Carol",
            "Modified at": dt.datetime(2025, 9, 27, 9, 0, 0),
            "Modified by": "Carol",
            "Steps": "Just one step with no markers",
            "Expected result": "Page renders",
            "Type": "ACCEPTANCE",
            "Priority": 1,
        },
        {
            "Folder": "Demo Project > Module B",
            "Test case ID": 2002,
            "Title": "Compatibility check",
            "Owner": "Dave",
            "Created at": dt.datetime(2025, 9, 27, 10, 0, 0),
            "Created by": "Dave",
            "Modified at": dt.datetime(2025, 9, 27, 10, 0, 0),
            "Modified by": "Dave",
            "Steps": "[1] open in browser X",
            "Expected result": "[1] no errors",
            "Type": "Compatibility",
            "Priority": 2,
            "Requirements": '[{"key":"DEMO-1","summary":"x","url":""},{"key":"DEMO-2"}]',
        },
        {
            "Folder": "Demo Project > Module B",
            "Test case ID": 2003,
            "Title": "Other type case",
            "Owner": "Eve",
            "Created at": dt.datetime(2025, 9, 27, 11, 0, 0),
            "Created by": "Eve",
            "Modified at": dt.datetime(2025, 9, 27, 11, 0, 0),
            "Modified by": "Eve",
            "Steps": "x",
            "Expected result": "y",
            "Type": "OTHER",
            "Priority": 3,
        },
    ])

    # Deprecated cases — should land with status="deprecated"
    add_sheet(wb, "Deprecated", [
        {
            "Folder": "DEPRECATED > Old Module",
            "Test case ID": 9001,
            "Title": "[DEPRECATED] Old test",
            "Owner": "Frank",
            "Created at": dt.datetime(2024, 1, 1, 0, 0, 0),
            "Created by": "Frank",
            "Modified at": dt.datetime(2024, 1, 1, 0, 0, 0),
            "Modified by": "Frank",
            "Steps": "[1] Old steps",
            "Expected result": "[1] Old expected",
            "Type": "FUNCTIONAL",
        },
        # row that should be SKIPPED for missing title
        {
            "Folder": "DEPRECATED",
            "Test case ID": 9002,
            "Title": None,
            "Owner": "Frank",
            "Created at": dt.datetime(2024, 1, 1, 0, 0, 0),
            "Created by": "Frank",
            "Modified at": dt.datetime(2024, 1, 1, 0, 0, 0),
            "Modified by": "Frank",
            "Steps": "[1] orphan steps",
            "Expected result": "[1] orphan expected",
            "Type": "FUNCTIONAL",
        },
    ])

    # CRITICAL CHANGES — Testiny's cross-cut root for migration suites
    add_sheet(wb, "Critical", [
        {
            "Folder": "CRITICAL CHANGES > Upgrade Suite",
            "Test case ID": 3001,
            "Title": "Verify upgrade smoke",
            "Owner": "Gina",
            "Created at": dt.datetime(2025, 3, 1, 0, 0, 0),
            "Created by": "Gina",
            "Modified at": dt.datetime(2025, 3, 1, 0, 0, 0),
            "Modified by": "Gina",
            "Steps": "[1] navigate\n[2] click",
            "Expected result": "[1] ok\n[2] ok",
            "Type": "Smoke & Sanity",
        },
    ])

    wb.save("fixture.xlsx")
    print("wrote fixture.xlsx")


if __name__ == "__main__":
    main()
